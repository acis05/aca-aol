from __future__ import annotations

import io
import math
import os
import re
import secrets
import time
import uuid
from datetime import datetime
from typing import Any, Dict, List, Tuple
from urllib.parse import urlencode

import pandas as pd
import requests
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, UploadFile, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.sessions import SessionMiddleware

load_dotenv()

# =========================
# ENV / CORE SETTINGS
# =========================
APP_ENV = os.getenv("APP_ENV", "dev").lower()  # dev|prod
DEBUG = os.getenv("DEBUG", "false").lower() == "true"

SESSION_SECRET = os.getenv("SESSION_SECRET", "CHANGE_ME_SUPER_SECRET")
HTTPS_ONLY = os.getenv("HTTPS_ONLY", "true").lower() == "true"  # MUST true in prod

# IMPORTANT:
# Untuk arsitektur login.aca-aol.id / abc.aca-aol.id,
# paling stabil pakai host-only cookie => COOKIE_DOMAIN kosong.
COOKIE_DOMAIN = os.getenv("COOKIE_DOMAIN", "").strip()  # kosongkan dulu di prod
COOKIE_SAMESITE = os.getenv("COOKIE_SAMESITE", "lax").lower()  # lax recommended

# Access code
DEFAULT_ACCESS_CODE = os.getenv("DEFAULT_ACCESS_CODE", "DEMO-ACA-001").strip()

TENANT_BASE_DOMAIN = os.getenv("TENANT_BASE_DOMAIN", "aca-aol.id").strip().lower()
TENANT_ACCESS_CODES = os.getenv("TENANT_ACCESS_CODES", "").strip()
STRICT_TENANT = os.getenv("STRICT_TENANT", "false").lower() == "true"

TIMEOUT_SEC = int(os.getenv("TIMEOUT_SEC", "60"))

# =========================
# OAUTH (ACCURATE)
# =========================
OAUTH_CLIENT_ID = os.getenv("OAUTH_CLIENT_ID", "")
OAUTH_CLIENT_SECRET = os.getenv("OAUTH_CLIENT_SECRET", "")
OAUTH_REDIRECT_URI = os.getenv("OAUTH_REDIRECT_URI", "")  # MUST be https://login.aca-aol.id/oauth/callback in prod
OAUTH_AUTH_URL = os.getenv("OAUTH_AUTH_URL", "https://account.accurate.id/oauth/authorize")
OAUTH_TOKEN_URL = os.getenv("OAUTH_TOKEN_URL", "https://account.accurate.id/oauth/token")
OAUTH_SCOPES = os.getenv("OAUTH_SCOPES", "journal_voucher_save sales_invoice_save sales_receipt_save")

ACCURATE_ACCOUNT_BASE_URL = os.getenv("ACCURATE_ACCOUNT_BASE_URL", "https://account.accurate.id").rstrip("/")
OPEN_DB_PATH = os.getenv("OPEN_DB_PATH", "/api/open-db.do")
DB_LIST_PATH = os.getenv("DB_LIST_PATH", "/api/db-list.do")

# Transaction endpoints (host from open-db)
JOURNAL_BULK_SAVE_PATH = os.getenv("JOURNAL_BULK_SAVE_PATH", "/accurate/api/journal-voucher/bulk-save.do")
SALES_INVOICE_BULK_SAVE_PATH = os.getenv("SALES_INVOICE_BULK_SAVE_PATH", "/accurate/api/sales-invoice/bulk-save.do")
SALES_RECEIPT_BULK_SAVE_PATH = os.getenv("SALES_RECEIPT_BULK_SAVE_PATH", "/accurate/api/sales-receipt/bulk-save.do")

# =========================
# APP SETUP
# =========================
app = FastAPI(title="ACA-AOL Accurate Importer")
templates = Jinja2Templates(directory="templates")

# In-memory job store (MVP)
JOBS: Dict[str, Dict[str, Any]] = {}

# fallback store
COMPANY_STORE = {"items": [{"id": "1161648", "name": "ACIS"}]}

# OAuth store (MVP single token)
OAUTH_STORE = {"state": None, "token": None, "token_expiry": 0}


# =========================
# UTIL: TENANT + HOST
# =========================
def _parse_tenant_codes() -> Dict[str, str]:
    """
    TENANT_ACCESS_CODES="login:KODELOGIN;abc:KODEABC"
    Return dict slug -> CODE_UPPER
    """
    out: Dict[str, str] = {}
    if not TENANT_ACCESS_CODES:
        return out
    parts = [p.strip() for p in TENANT_ACCESS_CODES.split(";") if p.strip()]
    for p in parts:
        if ":" not in p:
            continue
        slug, code = p.split(":", 1)
        slug = slug.strip().lower()
        code = code.strip().upper()
        if slug and code:
            out[slug] = code
    return out


TENANT_CODES = _parse_tenant_codes()


def get_host(request: Request) -> str:
    # Railway biasanya kirim host normal, tapi bisa juga via x-forwarded-host
    host = (request.headers.get("x-forwarded-host") or request.headers.get("host") or "").lower()
    return host.split(":")[0].strip()


def get_tenant_slug_from_request(request: Request) -> str:
    return "default"

    host = get_host(request)
    if not host:
        return "default"

    if host in ("localhost", "127.0.0.1"):
        return "default"

    if TENANT_BASE_DOMAIN and host.endswith(TENANT_BASE_DOMAIN):
        left = host[: -len(TENANT_BASE_DOMAIN)].rstrip(".")
        slug = left.split(".")[0].strip().lower() if left else "default"
        return slug or "default"

    return "default"


def is_public_path(path: str) -> bool:
    if path in ("/", "/favicon.ico"):
        return True
    if path.startswith("/auth/"):
        return True
    if path.startswith("/oauth/"):
        return True
    if path.startswith("/templates/"):
        return True
    if path.startswith("/health"):
        return True
    return False


def is_protected_path(path: str) -> bool:
    # semua API import + accurate kita proteksi
    return path.startswith("/api/") or path.startswith("/accurate/")


# =========================
# MIDDLEWARE: ACCESS GATE (needs session)
# =========================
class AccessGateMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path

        if is_public_path(path):
            return await call_next(request)

        if is_protected_path(path):
            if not request.session.get("access_ok"):
                return JSONResponse({"ok": False, "error": "Unauthorized. Masukkan kode akses dulu."}, status_code=401)

            if STRICT_TENANT:
                tenant = get_tenant_slug_from_request(request)
                sess_tenant = (request.session.get("tenant") or "default").lower()
                if tenant != sess_tenant:
                    return JSONResponse(
                        {"ok": False, "error": "Tenant session tidak cocok. Silakan login ulang."},
                        status_code=401,
                    )

        return await call_next(request)


app.add_middleware(AccessGateMiddleware)

# =========================
# SESSION MIDDLEWARE (MUST be outermost => add LAST)
# =========================
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    same_site="lax",
    https_only=True,
    domain="login.aca-aol.id",
)

app.add_middleware(AccessGateMiddleware)


# =========================
# GLOBAL ERROR HANDLER (biar gak "internal server error" gelap)
# =========================
@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    # jangan bocorin detail kalau prod
    if APP_ENV == "prod" and not DEBUG:
        return JSONResponse({"ok": False, "error": "Internal server error"}, status_code=500)

    return JSONResponse(
        {"ok": False, "error": f"{type(exc).__name__}: {str(exc)}", "path": request.url.path},
        status_code=500,
    )


# =========================
# AUTH ROUTES (ACCESS CODE)
# =========================
class VerifyAccessBody(BaseModel):
    code: str


@app.get("/auth/status")
def auth_status(request: Request):
    tenant = get_tenant_slug_from_request(request)

    configured = True
    if TENANT_CODES:
        # kalau kamu pakai subdomain codes, tenant harus ada di list
        configured = tenant in TENANT_CODES

    return {
        "ok": True,
        "tenant": tenant,
        "host": get_host(request),
        "configured": configured,
        "access_ok": bool(request.session.get("access_ok")),
        "session_tenant": request.session.get("tenant"),
        "https_only": HTTPS_ONLY if APP_ENV == "prod" else False,
        "cookie_domain": COOKIE_DOMAIN or None,
        "strict_tenant": STRICT_TENANT,
    }


@app.post("/auth/verify")
def auth_verify(request: Request, body: VerifyAccessBody):
    tenant = get_tenant_slug_from_request(request)
    code = (body.code or "").strip().upper()
    if not code:
        return JSONResponse({"ok": False, "error": "Kode akses kosong."}, status_code=400)

    # kalau TENANT_CODES diisi => per-subdomain
    if TENANT_CODES:
        expected = TENANT_CODES.get(tenant)
        if not expected:
            return JSONResponse({"ok": False, "error": f"Tenant '{tenant}' belum terdaftar."}, status_code=403)
        if not secrets.compare_digest(code, expected):
            return JSONResponse({"ok": False, "error": "Kode akses salah."}, status_code=401)
    else:
        # single code mode
        if not secrets.compare_digest(code, DEFAULT_ACCESS_CODE.upper()):
            return JSONResponse({"ok": False, "error": "Kode akses salah."}, status_code=401)

    request.session["access_ok"] = True
    request.session["tenant"] = tenant
    request.session["access_at"] = int(time.time())

    return {"ok": True, "tenant": tenant}


@app.post("/auth/logout")
def auth_logout(request: Request):
    request.session.clear()
    return {"ok": True}


@app.get("/auth/logout")
def auth_logout_get(request: Request):
    request.session.clear()
    return RedirectResponse("/", status_code=302)


# =========================
# UI ROUTES (simple)
# =========================
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    # Sederhana saja: halaman input access code + tombol oauth login (kamu sudah punya index.html)
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/app", response_class=HTMLResponse)
def app_home(request: Request):
    if not request.session.get("access_ok"):
        return RedirectResponse("/", status_code=302)
    # Kamu bisa pakai app.html sederhana
    return templates.TemplateResponse("app.html", {"request": request})


# =========================
# OAUTH HELPERS
# =========================
def oauth_build_auth_url() -> str:
    state = secrets.token_urlsafe(24)
    OAUTH_STORE["state"] = state
    params = {
        "response_type": "code",
        "client_id": OAUTH_CLIENT_ID,
        "redirect_uri": OAUTH_REDIRECT_URI,
        "scope": OAUTH_SCOPES,
        "state": state,
    }
    return f"{OAUTH_AUTH_URL}?{urlencode(params)}"


def oauth_exchange_code(code: str) -> dict:
    data = {"grant_type": "authorization_code", "code": code, "redirect_uri": OAUTH_REDIRECT_URI}
    r = requests.post(
        OAUTH_TOKEN_URL,
        data=data,
        auth=(OAUTH_CLIENT_ID, OAUTH_CLIENT_SECRET),
        timeout=30,
    )
    try:
        r.raise_for_status()
    except Exception:
        raise RuntimeError(f"{r.status_code} {r.text}")

    token = r.json()
    expires_in = int(token.get("expires_in", 3600))
    OAUTH_STORE["token"] = token
    OAUTH_STORE["token_expiry"] = int(time.time()) + max(60, expires_in - 60)
    return token


def get_access_token() -> str:
    token = OAUTH_STORE.get("token") or {}
    access = token.get("access_token")
    if not access:
        raise RuntimeError("Belum login OAuth / access_token belum ada. Klik Login Accurate dulu.")
    return access


# =========================
# OAUTH ROUTES
# =========================
@app.get("/oauth/login")
def oauth_login():
    if not (OAUTH_CLIENT_ID and OAUTH_CLIENT_SECRET and OAUTH_AUTH_URL and OAUTH_TOKEN_URL and OAUTH_REDIRECT_URI):
        return HTMLResponse("Config OAuth belum lengkap. Cek file .env (OAUTH_*).", status_code=500)
    return RedirectResponse(oauth_build_auth_url())


@app.get("/oauth/callback")
def oauth_callback(code: str = "", state: str = ""):
    if not code:
        return HTMLResponse("Tidak ada code dari Accurate. Login dibatalkan/gagal.", status_code=400)
    if not state or state != OAUTH_STORE.get("state"):
        return HTMLResponse("State tidak cocok. Silakan coba login ulang.", status_code=400)

    oauth_exchange_code(code)
    return RedirectResponse("/app", status_code=302)


@app.get("/oauth/status")
def oauth_status():
    token = OAUTH_STORE.get("token")
    if not token:
        return {"logged_in": False}
    return {
        "logged_in": True,
        "has_access_token": bool(token.get("access_token")),
        "token_type": token.get("token_type"),
        "scope": token.get("scope"),
        "token_expiry": OAUTH_STORE.get("token_expiry", 0),
    }


# =========================
# COMPANY PICKER API
# =========================
class SelectCompanyBody(BaseModel):
    id: str


class AddCompanyFromUrlBody(BaseModel):
    url: str


def fetch_companies_from_accurate() -> List[Dict[str, str]]:
    access = get_access_token()
    url = f"{ACCURATE_ACCOUNT_BASE_URL}{DB_LIST_PATH}"

    r = requests.get(url, headers={"Authorization": f"Bearer {access}", "Accept": "application/json"}, timeout=30)
    ct = (r.headers.get("content-type") or "").lower()
    if "application/json" not in ct:
        snippet = (r.text or "")[:300].replace("\n", " ").replace("\r", " ")
        raise RuntimeError(f"db-list non-JSON: status={r.status_code}, ct={ct}, body={snippet}")

    data = r.json()
    if data.get("s") is False:
        raise RuntimeError(f"db-list gagal: {data}")

    out: List[Dict[str, str]] = []
    for it in (data.get("d") or []):
        cid = str(it.get("id", "")).strip()
        name = str(it.get("alias") or it.get("name") or f"DB-{cid}").strip()
        if cid:
            out.append({"id": cid, "name": name})
    return out


def open_db_get_session_and_host(request: Request) -> Tuple[str, str]:
    """
    open-db ambil session + host.
    host wajib dipakai untuk /accurate/api/...
    """
    access = get_access_token()

    cid = request.session.get("selected_company_id")
    if not cid:
        raise RuntimeError("Belum pilih Data Usaha. Pilih dulu di dropdown.")

    url = f"{ACCURATE_ACCOUNT_BASE_URL}{OPEN_DB_PATH}"
    params = {"id": cid}

    r = requests.post(
        url,
        params=params,
        headers={"Authorization": f"Bearer {access}", "Accept": "application/json"},
        timeout=30,
        allow_redirects=False,
    )

    if r.status_code in (301, 302, 307, 308):
        raise RuntimeError(f"open-db redirect: {r.status_code}, location={r.headers.get('Location')}")

    ct = (r.headers.get("content-type") or "").lower()
    if "application/json" not in ct:
        snippet = (r.text or "")[:300].replace("\n", " ").replace("\r", " ")
        raise RuntimeError(f"open-db non-JSON: status={r.status_code}, ct={ct}, body={snippet}")

    data = r.json()
    if data.get("s") is False:
        raise RuntimeError(f"open-db gagal: {data}")

    sid = data.get("session") or data.get("sessionId") or data.get("xSessionId") or data.get("d")
    if not sid or isinstance(sid, (list, dict)):
        raise RuntimeError(f"Tidak ada session di response open-db: {data}")

    host = data.get("host") or data.get("h")
    if not host or not isinstance(host, str):
        raise RuntimeError(f"Tidak ada host di response open-db: {data}")

    host = host.rstrip("/")

    request.session["accurate_api_host"] = host
    request.session["x_session_id"] = sid

    return sid, host


def get_cached_session_and_host(request: Request) -> Tuple[str, str]:
    sid = (request.session.get("x_session_id") or "").strip()
    host = (request.session.get("accurate_api_host") or "").strip().rstrip("/")
    return sid, host


@app.get("/accurate/companies")
def accurate_companies(request: Request):
    selected = request.session.get("selected_company_id")
    try:
        token = (OAUTH_STORE.get("token") or {}).get("access_token")
        if token:
            items = fetch_companies_from_accurate()
            COMPANY_STORE["items"] = items
            return {"items": items, "selected_company_id": selected}
    except Exception as e:
        return {"items": COMPANY_STORE["items"], "selected_company_id": selected, "warning": str(e)}

    return {"items": COMPANY_STORE["items"], "selected_company_id": selected}


@app.post("/accurate/select-company")
def accurate_select_company(request: Request, body: SelectCompanyBody):
    cid = body.id.strip()
    if not cid.isdigit():
        return JSONResponse({"ok": False, "error": "ID harus angka"}, status_code=400)
    request.session["selected_company_id"] = cid
    request.session.pop("x_session_id", None)
    request.session.pop("accurate_api_host", None)
    return {"ok": True, "selected": cid}


@app.post("/accurate/add-company-from-url")
def accurate_add_company_from_url(request: Request, body: AddCompanyFromUrlBody):
    url = body.url.strip()
    m = re.search(r"[?&]id=(\d+)", url)
    if not m:
        return JSONResponse({"ok": False, "error": "URL tidak mengandung id=angka. Contoh: ...open.do?id=1161648"}, status_code=400)
    cid = m.group(1)

    for it in COMPANY_STORE["items"]:
        if it["id"] == cid:
            return {"ok": True, "item": it}

    item = {"id": cid, "name": f"DataUsaha-{cid}"}
    COMPANY_STORE["items"].append(item)
    return {"ok": True, "item": item}


# =========================
# EXCEL HELPERS + NORMALIZATION
# =========================
def is_nan(x: Any) -> bool:
    return isinstance(x, float) and math.isnan(x)


def normalize_str(x: Any) -> str:
    if x is None or is_nan(x):
        return ""
    return str(x).strip()


def normalize_code(x: Any) -> str:
    if isinstance(x, float) and x.is_integer():
        return str(int(x))
    return normalize_str(x)


def normalize_int_str(x: Any) -> str:
    if x is None or is_nan(x):
        return ""
    if isinstance(x, float) and x.is_integer():
        return str(int(x))
    s = normalize_str(x)
    return s if s.isdigit() else ""


def parse_number(raw: Any) -> Tuple[bool, float, str]:
    if raw is None or is_nan(raw):
        return False, 0.0, "nilai kosong"
    try:
        if isinstance(raw, (int, float)):
            return True, float(raw), ""
        s = normalize_str(raw)
        if not s:
            return False, 0.0, "nilai kosong"
        s = s.replace(" ", "")
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        val = float(s)
        return True, val, ""
    except Exception:
        return False, 0.0, f"nilai bukan angka: {raw}"


def parse_bool(raw: Any) -> str:
    s = normalize_str(raw).strip().lower()
    if not s:
        return ""
    if s in ("1", "true", "yes", "y"):
        return "true"
    if s in ("0", "false", "no", "n"):
        return "false"
    return ""


def safe_money(raw: Any) -> Tuple[bool, float, str]:
    ok, val, err = parse_number(raw)
    if not ok:
        return False, 0.0, err.replace("nilai", "amount")
    if val <= 0:
        return False, val, "amount harus > 0"
    return True, val, ""


def safe_qty(raw: Any) -> Tuple[bool, float, str]:
    if raw is None or is_nan(raw) or normalize_str(raw) == "":
        return True, 1.0, ""
    ok, val, err = parse_number(raw)
    if not ok:
        return False, 0.0, err.replace("nilai", "quantity")
    if val <= 0:
        return False, val, "quantity harus > 0"
    return True, val, ""


def parse_date_to_ddmmyyyy(raw: Any) -> Tuple[bool, str]:
    if raw is None or is_nan(raw):
        return False, "Tanggal kosong"
    if isinstance(raw, datetime):
        return True, raw.strftime("%d/%m/%Y")

    s = normalize_str(raw)
    if not s:
        return False, "Tanggal kosong"

    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        return True, s

    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            dt = datetime.strptime(s, "%Y-%m-%d")
            return True, dt.strftime("%d/%m/%Y")
        except Exception:
            return False, f"Format tanggal tidak valid: {s}"

    if re.match(r"^\d{2}-\d{2}-\d{4}$", s):
        try:
            dt = datetime.strptime(s, "%d-%m-%Y")
            return True, dt.strftime("%d/%m/%Y")
        except Exception:
            return False, f"Format tanggal tidak valid: {s}"

    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return True, dt.strftime("%d/%m/%Y")
        except Exception:
            pass

    return False, f"Format tanggal tidak dikenali: {s}"


def read_excel_to_df(file_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or not rows[0]:
        raise ValueError("Sheet kosong")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not any(headers):
        raise ValueError("Header kosong")

    data_rows = []
    for r in rows[1:]:
        if r is None:
            continue
        row_dict = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            val = r[i] if i < len(r) else None
            if val not in (None, ""):
                empty = False
            row_dict[h] = val
        if not empty:
            data_rows.append(row_dict)

    return pd.DataFrame(data_rows)


def chunk_list(items: List[Any], size: int) -> List[List[Any]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


# =========================
# JOURNAL VOUCHER BUILDER (dipertahankan dari versi kamu)
# =========================
REQUIRED_COLS = ["jurnal_key", "transDate", "accountNo", "amountType", "amount"]
AMOUNTTYPE_ALLOWED = {"DEBIT", "CREDIT"}
SUBSIDIARY_ALLOWED = {"CUSTOMER", "VENDOR", "EMPLOYEE"}
AMOUNTTYPE_MAP = {
    "D": "DEBIT", "DEBIT": "DEBIT", "DR": "DEBIT",
    "K": "CREDIT", "KREDIT": "CREDIT", "CREDIT": "CREDIT", "CR": "CREDIT",
}


def get_first_nonempty(row, keys: List[str]) -> str:
    for k in keys:
        v = normalize_str(row.get(k))
        if v:
            return v
    return ""


def normalize_amount_type(raw: Any) -> Tuple[bool, str]:
    s = normalize_str(raw).upper()
    if not s:
        return False, "amountType kosong"
    s = AMOUNTTYPE_MAP.get(s, s)
    if s not in AMOUNTTYPE_ALLOWED:
        return False, f"amountType '{raw}' tidak valid (pakai DEBIT/CREDIT)"
    return True, s


def validate_subsidiary(detail: Dict[str, Any]) -> List[str]:
    errs: List[str] = []
    st = normalize_str(detail.get("subsidiaryType", "")).upper()
    if not st:
        return errs
    if st not in SUBSIDIARY_ALLOWED:
        errs.append(f"subsidiaryType '{st}' tidak valid (CUSTOMER/VENDOR/EMPLOYEE)")
        return errs

    if st == "CUSTOMER" and not normalize_str(detail.get("customerNo")):
        errs.append("subsidiaryType=CUSTOMER tapi customerNo kosong")
    if st == "VENDOR" and not normalize_str(detail.get("vendorNo")):
        errs.append("subsidiaryType=VENDOR tapi vendorNo kosong")
    if st == "EMPLOYEE" and not normalize_str(detail.get("employeeNo")):
        errs.append("subsidiaryType=EMPLOYEE tapi employeeNo kosong")
    return errs


def build_journals_from_df(df: pd.DataFrame) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    errors: List[Dict[str, Any]] = []

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        for c in missing:
            errors.append({"type": "missing_column", "jurnal_key": "", "row": None, "message": f"Kolom wajib tidak ada: {c}"})
        return [], errors

    parsed_rows: List[Dict[str, Any]] = []

    for idx, row in df.iterrows():
        row_num = int(idx) + 2
        jurnal_key = normalize_str(row.get("jurnal_key"))
        if not jurnal_key:
            errors.append({"type": "row_error", "jurnal_key": "", "row": row_num, "message": "jurnal_key kosong"})
            continue

        ok_date, ddmmyyyy_or_err = parse_date_to_ddmmyyyy(row.get("transDate"))
        if not ok_date:
            errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": ddmmyyyy_or_err})
            continue

        account_no = normalize_str(row.get("accountNo"))
        if not account_no:
            errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": "accountNo kosong"})
            continue

        ok_at, at_or_err = normalize_amount_type(row.get("amountType"))
        if not ok_at:
            errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": at_or_err})
            continue

        ok_amt, amt, amt_err = safe_money(row.get("amount"))
        if not ok_amt:
            errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": amt_err})
            continue

        rate_raw_present = normalize_str(row.get("rate")) != ""
        prime_raw_present = normalize_str(row.get("primeAmount")) != ""

        rate_str = ""
        prime_str = ""

        if rate_raw_present:
            ok_rate, rate_val, rate_err = parse_number(row.get("rate"))
            if (not ok_rate) or rate_val <= 0:
                errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": f"rate tidak valid: {rate_err}"})
                continue
            rate_str = f"{rate_val:.6f}".rstrip("0").rstrip(".")

        if prime_raw_present:
            ok_prime, prime_val, prime_err = parse_number(row.get("primeAmount"))
            if (not ok_prime) or prime_val <= 0:
                errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": f"primeAmount tidak valid: {prime_err}"})
                continue
            prime_str = f"{prime_val:.6f}".rstrip("0").rstrip(".")

        if (rate_raw_present and not prime_raw_present) or (prime_raw_present and not rate_raw_present):
            errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": "Untuk mata uang asing, isi rate dan primeAmount (keduanya)."})
            continue

        header = {
            "transDate": ddmmyyyy_or_err,
            "description": normalize_str(row.get("description")),
            "number": get_first_nonempty(row, ["number", "journalNumber", "voucherNo", "voucherNo.", "Voucher No", "Nomor", "nomor"]),
            "branchId": normalize_int_str(row.get("branchId")),
            "branchName": normalize_str(row.get("branchName")),
            "typeAutoNumber": normalize_int_str(row.get("typeAutoNumber")),
        }

        detail: Dict[str, Any] = {
            "accountNo": account_no,
            "amountType": at_or_err,
            "amount": amt,
            "memo": normalize_str(row.get("memo")),
            "subsidiaryType": normalize_str(row.get("subsidiaryType")).upper(),
            "customerNo": normalize_code(row.get("customerNo")),
            "vendorNo": normalize_code(row.get("vendorNo")),
            "employeeNo": normalize_code(row.get("employeeNo")),
            "departmentName": normalize_str(row.get("departmentName")),
            "projectNo": normalize_str(row.get("projectNo")),
            "rate": rate_str,
            "primeAmount": prime_str,
        }

        for i in range(1, 11):
            k = f"dataClassification{i}Name"
            if k in df.columns:
                detail[k] = normalize_str(row.get(k))

        sub_errs = validate_subsidiary(detail)
        if sub_errs:
            for e in sub_errs:
                errors.append({"type": "row_error", "jurnal_key": jurnal_key, "row": row_num, "message": e})
            continue

        parsed_rows.append({"jurnal_key": jurnal_key, "row_num": row_num, "header": header, "detail": detail})

    if not parsed_rows:
        return [], errors

    by_key: Dict[str, List[Dict[str, Any]]] = {}
    for r in parsed_rows:
        by_key.setdefault(r["jurnal_key"], []).append(r)

    journals: List[Dict[str, Any]] = []

    for k, rows in by_key.items():
        first_header = rows[0]["header"]
        trans_date = first_header["transDate"]

        mismatch = [r for r in rows if r["header"]["transDate"] != trans_date]
        if mismatch:
            errors.append({"type": "journal_error", "jurnal_key": k, "row": None, "message": "transDate berbeda-beda dalam jurnal yang sama"})
            continue

        details = [r["detail"] for r in rows]
        if len(details) < 2:
            errors.append({"type": "journal_error", "jurnal_key": k, "row": None, "message": "Jurnal minimal 2 baris detail"})
            continue

        debit_total = sum(d["amount"] for d in details if d["amountType"] == "DEBIT")
        credit_total = sum(d["amount"] for d in details if d["amountType"] == "CREDIT")

        if abs(debit_total - credit_total) > 0.000001:
            errors.append({"type": "journal_error", "jurnal_key": k, "row": None, "message": f"Jurnal tidak balance (DEBIT={debit_total:g}, CREDIT={credit_total:g})"})
            continue

        header_payload: Dict[str, Any] = {"transDate": trans_date}
        for field in ["description", "number", "branchId", "branchName", "typeAutoNumber", "id"]:
            val = normalize_str(first_header.get(field))
            if val:
                header_payload[field] = val

        journals.append({"jurnal_key": k, "header": header_payload, "details": details, "totals": {"debit": debit_total, "credit": credit_total}})

    return journals, errors


def journals_to_form_payload(journals: List[Dict[str, Any]]) -> Dict[str, str]:
    payload: Dict[str, str] = {}

    for i, j in enumerate(journals):
        header = j["header"]
        payload[f"data[{i}].transDate"] = str(header["transDate"])

        for field in ["description", "number", "branchId", "branchName", "typeAutoNumber", "id"]:
            if field in header and normalize_str(header[field]):
                payload[f"data[{i}].{field}"] = normalize_str(header[field])

        for m, d in enumerate(j["details"]):
            base = f"data[{i}].detailJournalVoucher[{m}]"
            payload[f"{base}.accountNo"] = normalize_str(d.get("accountNo"))
            payload[f"{base}.amount"] = f"{float(d.get('amount')):.6f}".rstrip("0").rstrip(".")
            payload[f"{base}.amountType"] = normalize_str(d.get("amountType")).upper()

            for field in [
                "memo", "subsidiaryType", "customerNo", "vendorNo", "employeeNo",
                "departmentName", "projectNo", "rate", "primeAmount", "_status", "id",
            ] + [f"dataClassification{x}Name" for x in range(1, 11)]:
                if field in d and normalize_str(d[field]):
                    payload[f"{base}.{field}"] = normalize_str(d[field])

    return payload


# =========================
# SALES INVOICE BUILDER (dipertahankan dari versi kamu)
# =========================
SALESINV_REQUIRED_COLS = ["invoice_key", "customerNo", "itemNo", "unitPrice"]


def build_sales_invoices_from_df(df: pd.DataFrame) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    errors: List[Dict[str, Any]] = []

    missing = [c for c in SALESINV_REQUIRED_COLS if c not in df.columns]
    if missing:
        for c in missing:
            errors.append({"type": "missing_column", "invoice_key": "", "row": None, "message": f"Kolom wajib tidak ada: {c}"})
        return [], errors

    parsed_rows: List[Dict[str, Any]] = []

    for idx, row in df.iterrows():
        row_num = int(idx) + 2
        inv_key = normalize_str(row.get("invoice_key"))
        if not inv_key:
            errors.append({"type": "row_error", "invoice_key": "", "row": row_num, "message": "invoice_key kosong"})
            continue

        customer_no = normalize_code(row.get("customerNo"))
        if not customer_no:
            errors.append({"type": "row_error", "invoice_key": inv_key, "row": row_num, "message": "customerNo kosong"})
            continue

        trans_raw = row.get("transDate")
        trans_date = ""
        if normalize_str(trans_raw):
            ok_date, ddmmyyyy_or_err = parse_date_to_ddmmyyyy(trans_raw)
            if not ok_date:
                errors.append({"type": "row_error", "invoice_key": inv_key, "row": row_num, "message": ddmmyyyy_or_err})
                continue
            trans_date = ddmmyyyy_or_err

        item_no = normalize_code(row.get("itemNo"))
        if not item_no:
            errors.append({"type": "row_error", "invoice_key": inv_key, "row": row_num, "message": "itemNo kosong"})
            continue

        ok_price, price, price_err = parse_number(row.get("unitPrice"))
        if (not ok_price) or price <= 0:
            errors.append({"type": "row_error", "invoice_key": inv_key, "row": row_num, "message": f"unitPrice tidak valid: {price_err}"})
            continue

        ok_qty, qty, qty_err = safe_qty(row.get("quantity"))
        if not ok_qty:
            errors.append({"type": "row_error", "invoice_key": inv_key, "row": row_num, "message": qty_err})
            continue

        header = {
            "customerNo": customer_no,
            "transDate": trans_date,
            "number": normalize_str(row.get("number")),
            "description": normalize_str(row.get("description")),
            "currencyCode": normalize_str(row.get("currencyCode")).upper(),
            "rate": normalize_str(row.get("rate")),
            "fiscalRate": normalize_str(row.get("fiscalRate")),
            "inclusiveTax": parse_bool(row.get("inclusiveTax")),
            "taxable": parse_bool(row.get("taxable")),
            "taxType": normalize_str(row.get("taxType")),
            "documentCode": normalize_str(row.get("documentCode")),
            "documentTransaction": normalize_str(row.get("documentTransaction")),
            "shipDate": normalize_str(row.get("shipDate")),
            "shipmentName": normalize_str(row.get("shipmentName")),
            "paymentTermName": normalize_str(row.get("paymentTermName")),
            "poNumber": normalize_str(row.get("poNumber")),
            "toAddress": normalize_str(row.get("toAddress")),
            "typeAutoNumber": normalize_int_str(row.get("typeAutoNumber")),
        }

        detail: Dict[str, Any] = {
            "itemNo": item_no,
            "unitPrice": f"{price:.6f}".rstrip("0").rstrip("."),
            "quantity": f"{qty:.6f}".rstrip("0").rstrip("."),
            "warehouseName": normalize_str(row.get("warehouseName")),
            "itemUnitName": normalize_str(row.get("itemUnitName")),
            "detailName": normalize_str(row.get("detailName")),
            "detailNotes": normalize_str(row.get("detailNotes")),
            "itemDiscPercent": normalize_str(row.get("itemDiscPercent")),
            "itemCashDiscount": normalize_str(row.get("itemCashDiscount")),
            "projectNo": normalize_str(row.get("projectNo")),
            "departmentName": normalize_str(row.get("departmentName")),
            "deliveryOrderNumber": normalize_str(row.get("deliveryOrderNumber")),
            "salesOrderNumber": normalize_str(row.get("salesOrderNumber")),
            "salesQuotationNumber": normalize_str(row.get("salesQuotationNumber")),
        }

        for i in range(1, 11):
            k = f"dataClassification{i}Name"
            if k in df.columns:
                detail[k] = normalize_str(row.get(k))

        parsed_rows.append({"invoice_key": inv_key, "row_num": row_num, "header": header, "detail": detail})

    if not parsed_rows:
        return [], errors

    by_key: Dict[str, List[Dict[str, Any]]] = {}
    for r in parsed_rows:
        by_key.setdefault(r["invoice_key"], []).append(r)

    invoices: List[Dict[str, Any]] = []

    for k, rows in by_key.items():
        first_header = rows[0]["header"]
        td = normalize_str(first_header.get("transDate"))

        mismatch = [r for r in rows if normalize_str(r["header"].get("transDate")) not in ("", td)]
        if td and mismatch:
            errors.append({"type": "invoice_error", "invoice_key": k, "row": None, "message": "transDate berbeda-beda dalam invoice yang sama"})
            continue

        if not td:
            errors.append({"type": "invoice_error", "invoice_key": k, "row": None, "message": "transDate wajib diisi (dd/mm/yyyy)"})
            continue

        details = [r["detail"] for r in rows]
        if len(details) < 1:
            errors.append({"type": "invoice_error", "invoice_key": k, "row": None, "message": "Invoice minimal 1 baris item"})
            continue

        header_payload: Dict[str, Any] = {"customerNo": normalize_str(first_header.get("customerNo")), "transDate": td}
        for f in [
            "number","description","currencyCode","rate","fiscalRate","taxType",
            "documentCode","documentTransaction","shipDate","shipmentName",
            "paymentTermName","poNumber","toAddress","typeAutoNumber"
        ]:
            v = normalize_str(first_header.get(f))
            if v:
                header_payload[f] = v

        for bf in ["inclusiveTax", "taxable"]:
            v = normalize_str(first_header.get(bf))
            if v in ("true", "false"):
                header_payload[bf] = v

        invoices.append({"invoice_key": k, "header": header_payload, "detailItem": details})

    return invoices, errors


def sales_invoices_to_form_payload(invoices: List[Dict[str, Any]]) -> Dict[str, str]:
    payload: Dict[str, str] = {}

    for i, inv in enumerate(invoices):
        h = inv["header"]
        payload[f"data[{i}].customerNo"] = normalize_str(h.get("customerNo"))
        payload[f"data[{i}].transDate"] = normalize_str(h.get("transDate"))

        for field in [
            "number","description","currencyCode","rate","fiscalRate","taxType",
            "documentCode","documentTransaction","shipDate","shipmentName",
            "paymentTermName","poNumber","toAddress","typeAutoNumber",
        ]:
            v = normalize_str(h.get(field))
            if v:
                payload[f"data[{i}].{field}"] = v

        for bf in ["inclusiveTax", "taxable"]:
            v = normalize_str(h.get(bf))
            if v in ("true", "false"):
                payload[f"data[{i}].{bf}"] = v

        items = inv.get("detailItem") or []
        for m, d in enumerate(items):
            base = f"data[{i}].detailItem[{m}]"
            payload[f"{base}.itemNo"] = normalize_str(d.get("itemNo"))
            payload[f"{base}.unitPrice"] = normalize_str(d.get("unitPrice"))
            payload[f"{base}.quantity"] = normalize_str(d.get("quantity"))

            for field in [
                "warehouseName","itemUnitName","detailName","detailNotes",
                "itemDiscPercent","itemCashDiscount","projectNo","departmentName",
                "deliveryOrderNumber","salesOrderNumber","salesQuotationNumber",
            ] + [f"dataClassification{x}Name" for x in range(1, 11)]:
                v = normalize_str(d.get(field))
                if v:
                    payload[f"{base}.{field}"] = v

    return payload


# =========================
# POSTERS (bulk-save)
# =========================
def post_journal_bulk_save(request: Request, journals_batch: List[Dict[str, Any]]) -> Tuple[bool, Any]:
    access = get_access_token()

    sid, host = get_cached_session_and_host(request)
    if not sid or not host:
        sid, host = open_db_get_session_and_host(request)

    url = host.rstrip("/") + JOURNAL_BULK_SAVE_PATH
    headers = {"X-Session-ID": sid, "Authorization": f"Bearer {access}", "Accept": "application/json"}
    form_payload = journals_to_form_payload(journals_batch)

    resp = requests.post(url, headers=headers, data=form_payload, timeout=TIMEOUT_SEC, allow_redirects=False)

    try:
        data = resp.json()
    except Exception:
        data = resp.text

    if 200 <= resp.status_code < 300:
        return True, data
    return False, {"status_code": resp.status_code, "response": data, "url": url}


def post_sales_invoice_bulk_save(request: Request, invoices_batch: List[Dict[str, Any]]) -> Tuple[bool, Any]:
    access = get_access_token()

    sid, host = get_cached_session_and_host(request)
    if not sid or not host:
        sid, host = open_db_get_session_and_host(request)

    url = host.rstrip("/") + SALES_INVOICE_BULK_SAVE_PATH
    headers = {"X-Session-ID": sid, "Authorization": f"Bearer {access}", "Accept": "application/json"}
    form_payload = sales_invoices_to_form_payload(invoices_batch)

    resp = requests.post(url, headers=headers, data=form_payload, timeout=TIMEOUT_SEC, allow_redirects=False)

    try:
        data = resp.json()
    except Exception:
        data = resp.text

    if 200 <= resp.status_code < 300:
        return True, data
    return False, {"status_code": resp.status_code, "response": data, "url": url}


# =========================
# API: JOURNAL VOUCHER
# =========================
@app.post("/api/journal/preview")
async def api_journal_preview(file: UploadFile = File(...)):
    content = await file.read()
    df = read_excel_to_df(content)
    return {"ok": True, "columns": list(df.columns), "preview": df.head(15).fillna("").to_dict(orient="records")}


@app.post("/api/journal/validate")
async def api_journal_validate(file: UploadFile = File(...)):
    content = await file.read()
    df = read_excel_to_df(content)
    journals, errors = build_journals_from_df(df)

    return {
        "ok": True,
        "summary": {
            "total_rows": int(len(df)),
            "total_journals_detected": int(df["jurnal_key"].nunique()) if "jurnal_key" in df.columns else 0,
            "valid_journals": int(len(journals)),
            "errors": int(len(errors)),
        },
        "errors": errors[:500],
    }


@app.post("/api/journal/import")
async def api_journal_import(request: Request, file: UploadFile = File(...), session_id: str = Form("")):
    content = await file.read()
    df = read_excel_to_df(content)

    journals, errors = build_journals_from_df(df)
    if errors:
        return JSONResponse({"ok": False, "error": "Masih ada error validasi. Perbaiki dulu.", "errors": errors[:500]}, status_code=400)
    if not journals:
        return JSONResponse({"ok": False, "error": "Tidak ada jurnal valid untuk dikirim."}, status_code=400)

    if session_id.strip():
        request.session["x_session_id"] = session_id.strip()

    # ensure session/host
    sid, host = get_cached_session_and_host(request)
    if not sid or not host:
        open_db_get_session_and_host(request)

    job_id = str(uuid.uuid4())
    JOBS[job_id] = {"status": "running", "logs": [], "summary": {}}

    batches = chunk_list(journals, 100)
    success_count = 0
    fail_count = 0

    for bi, batch in enumerate(batches, start=1):
        ok, resp = post_journal_bulk_save(request=request, journals_batch=batch)
        entry: Dict[str, Any] = {"batch": bi, "sent": len(batch), "ok": ok, "response": resp}
        JOBS[job_id]["logs"].append(entry)

        if not ok:
            fail_count += len(batch)
            continue

        items: List[Any] = resp.get("d", []) if isinstance(resp, dict) else []
        for i, _j in enumerate(batch):
            item = items[i] if i < len(items) else None
            if isinstance(item, dict) and item.get("s") is True:
                success_count += 1
            else:
                fail_count += 1

    JOBS[job_id]["status"] = "done"
    JOBS[job_id]["summary"] = {
        "total_valid_journals": len(journals),
        "batches": len(batches),
        "success_count": success_count,
        "failed_count": fail_count,
        "used_session_id_prefix": (request.session.get("x_session_id") or "")[:8],
        "used_host": request.session.get("accurate_api_host") or "",
        "endpoint_path": JOURNAL_BULK_SAVE_PATH,
    }
    return {"ok": True, "job_id": job_id, "summary": JOBS[job_id]["summary"]}


# =========================
# API: SALES INVOICE
# =========================
@app.post("/api/sales-invoice/preview")
async def api_sales_invoice_preview(file: UploadFile = File(...)):
    content = await file.read()
    df = read_excel_to_df(content)
    return {"ok": True, "columns": list(df.columns), "preview": df.head(15).fillna("").to_dict(orient="records")}


@app.post("/api/sales-invoice/validate")
async def api_sales_invoice_validate(file: UploadFile = File(...)):
    content = await file.read()
    df = read_excel_to_df(content)
    invoices, errors = build_sales_invoices_from_df(df)

    return {
        "ok": True,
        "summary": {
            "total_rows": int(len(df)),
            "total_invoices_detected": int(df["invoice_key"].nunique()) if "invoice_key" in df.columns else 0,
            "valid_invoices": int(len(invoices)),
            "errors": int(len(errors)),
        },
        "errors": errors[:500],
    }


@app.post("/api/sales-invoice/import")
async def api_sales_invoice_import(request: Request, file: UploadFile = File(...)):
    content = await file.read()
    df = read_excel_to_df(content)

    invoices, errors = build_sales_invoices_from_df(df)
    if errors:
        return JSONResponse({"ok": False, "error": "Masih ada error validasi. Perbaiki dulu.", "errors": errors[:500]}, status_code=400)
    if not invoices:
        return JSONResponse({"ok": False, "error": "Tidak ada invoice valid untuk dikirim."}, status_code=400)

    sid, host = get_cached_session_and_host(request)
    if not sid or not host:
        open_db_get_session_and_host(request)

    job_id = str(uuid.uuid4())
    JOBS[job_id] = {"status": "running", "logs": [], "summary": {}}

    batches = chunk_list(invoices, 100)
    success_count = 0
    fail_count = 0

    for bi, batch in enumerate(batches, start=1):
        ok, resp = post_sales_invoice_bulk_save(request=request, invoices_batch=batch)
        entry: Dict[str, Any] = {"batch": bi, "sent": len(batch), "ok": ok, "response": resp}
        JOBS[job_id]["logs"].append(entry)

        if not ok:
            fail_count += len(batch)
            continue

        items: List[Any] = resp.get("d", []) if isinstance(resp, dict) else []
        for i, _inv in enumerate(batch):
            item = items[i] if i < len(items) else None
            if isinstance(item, dict) and item.get("s") is True:
                success_count += 1
            else:
                fail_count += 1

    JOBS[job_id]["status"] = "done"
    JOBS[job_id]["summary"] = {
        "total_valid_invoices": len(invoices),
        "batches": len(batches),
        "success_count": success_count,
        "failed_count": fail_count,
        "used_session_id_prefix": (request.session.get("x_session_id") or "")[:8],
        "used_host": request.session.get("accurate_api_host") or "",
        "endpoint_path": SALES_INVOICE_BULK_SAVE_PATH,
    }
    return {"ok": True, "job_id": job_id, "summary": JOBS[job_id]["summary"]}


# =========================
# API: SALES RECEIPT (SLOT)
# =========================
@app.post("/api/sales-receipt/preview")
async def api_sales_receipt_preview():
    return JSONResponse(
        {"ok": False, "error": "Sales Receipt belum ditempel logic builder/payload-nya di app.py ini."},
        status_code=501,
    )

@app.post("/api/sales-receipt/validate")
async def api_sales_receipt_validate():
    return JSONResponse(
        {"ok": False, "error": "Sales Receipt belum ditempel logic builder/payload-nya di app.py ini."},
        status_code=501,
    )

@app.post("/api/sales-receipt/import")
async def api_sales_receipt_import():
    return JSONResponse(
        {"ok": False, "error": "Sales Receipt belum ditempel logic builder/payload-nya di app.py ini."},
        status_code=501,
    )


# =========================
# JOB STATUS
# =========================
@app.get("/api/job/{job_id}")
def api_job(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        return JSONResponse({"ok": False, "error": "Job tidak ditemukan"}, status_code=404)
    return {"ok": True, "job": job}


# =========================
# TEMPLATES DOWNLOAD
# =========================
def _make_template_xlsx(headers: List[str], example_row: List[Any], sheet_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.append(example_row)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.get("/templates/journal-voucher.xlsx")
def download_template_journal_voucher():
    headers = [
        "jurnal_key", "transDate", "accountNo", "amountType", "amount",
        "description", "number", "memo",
        "subsidiaryType", "customerNo", "vendorNo", "employeeNo",
        "rate", "primeAmount",
    ]
    example = [
        "JV-001", "17/01/2026", "1101", "DEBIT", 100000,
        "Contoh jurnal", "JV/01/2026/001", "memo baris 1",
        "", "", "", "",
        "", "",
    ]
    content = _make_template_xlsx(headers, example, "JournalVoucher")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template-journal-voucher.xlsx"},
    )


@app.get("/templates/sales-invoice.xlsx")
def download_template_sales_invoice():
    headers = [
        "invoice_key", "transDate", "customerNo", "itemNo", "unitPrice", "quantity",
        "number", "description", "currencyCode", "rate",
        "inclusiveTax", "taxable", "taxType",
        "warehouseName", "itemUnitName",
        "detailName", "detailNotes",
        "itemDiscPercent", "itemCashDiscount",
    ]
    example = [
        "SI-001", "17/01/2026", "CUST-001", "ITEM-001", 50000, 2,
        "SI/01/2026/001", "Contoh faktur", "IDR", "",
        "false", "true", "",
        "Gudang Utama", "PCS",
        "Nama detail", "Catatan detail",
        "", "",
    ]
    content = _make_template_xlsx(headers, example, "SalesInvoice")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template-sales-invoice.xlsx"},
    )


# =========================
# HEALTH
# =========================
@app.get("/health")
def health(request: Request):
    return {
        "ok": True,
        "env": APP_ENV,
        "host": get_host(request),
        "tenant": get_tenant_slug_from_request(request),
        "access_ok": bool(request.session.get("access_ok")),
        "oauth_logged_in": bool((OAUTH_STORE.get("token") or {}).get("access_token")),
        "selected_company_id": request.session.get("selected_company_id"),
        "cookie_domain": COOKIE_DOMAIN or None,
        "https_only": HTTPS_ONLY if APP_ENV == "prod" else False,
        "strict_tenant": STRICT_TENANT,
    }
